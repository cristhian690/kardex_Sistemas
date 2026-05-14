import io 
import enum
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def exportar_kardex_excel(df: pd.DataFrame) -> bytes:
    """
    Genera un archivo Excel con el kardex procesado.
    Formato: cabeceras agrupadas, bordes, anchos optimizados.
    Retorna los bytes del archivo .xlsx listo para descargar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    # ── Estilos ────────────────────────────────────────────────────────────────
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin")
    borde  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Cabeceras agrupadas (fila 1) ───────────────────────────────────────────
    headers_grupo = [
        ("Código",            1, 1),
        ("COMPROBANTE",       2, 5),
        ("Tipo de Operación", 6, 6),
        ("ENTRADAS",          7, 9),
        ("SALIDAS",           10, 12),
        ("SALDO FINAL",       13, 15),
    ]

    for (titulo, col_ini, col_fin) in headers_grupo:
        cell = ws.cell(row=1, column=col_ini, value=titulo)
        cell.font      = bold
        cell.alignment = center
        cell.border    = borde
        if col_ini != col_fin:
            ws.merge_cells(
                start_row=1, start_column=col_ini,
                end_row=1,   end_column=col_fin
            )
        else:
            ws.merge_cells(
                start_row=1, start_column=col_ini,
                end_row=2,   end_column=col_fin
            )
            ws.cell(row=2, column=col_ini).border = borde

    # ── Subcabeceras (fila 2) ──────────────────────────────────────────────────
    subheaders = [
        "Código",
        "Fecha", "Tipo", "Serie", "Número", "Tipo de Operación",
        "Cantidad", "Costo Unitario", "Costo Total",
        "Cantidad", "Costo Unitario", "Costo Total",
        "Cantidad", "Costo Unitario", "Costo Total",
    ]
    for col, sh in enumerate(subheaders, start=1):
        if col in (1, 6):
            continue
        cell = ws.cell(row=2, column=col, value=sh)
        cell.font      = bold
        cell.alignment = center
        cell.border    = borde

    # ── Datos ──────────────────────────────────────────────────────────────────
    cols_centradas = {2, 3, 4, 5, 6}

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        numero_val = row.Numero
        try:
            numero_val = int(float(str(numero_val))) if str(numero_val).strip() not in ("", "nan") else 0
        except Exception:
            numero_val = 0

        fecha_val = ""
        try:
            if pd.notna(row.Fecha):
                fecha_val = pd.Timestamp(row.Fecha).strftime("%d/%m/%Y")
        except Exception:
            fecha_val = ""

        # Conversión al valor de Enum a valor para exportar excel
        tipo_operacion_val = row.Tipo_Operacion.value if isinstance(row.Tipo_Operacion, enum.Enum) else row.Tipo_Operacion

        datos = [
            (str(row.Codigo),      "@"),
            (fecha_val,            "@"),
            (row.Tipo,             "00"),
            (str(row.Serie),       "@"),
            (numero_val,           r'[$-408]00000000'),
            (tipo_operacion_val,   "@"),  # ya es cadena
            (row.Ent_Cantidad,     "#,##0.000"),
            (row.Ent_Costo_Unit,   "#,##0.0000"),
            (row.Ent_Costo_Total,  "#,##0.000"),
            (row.Sal_Cantidad,     "#,##0.000"),
            (row.Sal_Costo_Unit,   "#,##0.0000"),
            (row.Sal_Costo_Total,  "#,##0.000"),
            (row.Saldo_Cantidad,   "#,##0.000"),
            (row.Saldo_Costo_Unit, "#,##0.0000"),
            (row.Saldo_Costo_Total,"#,##0.000"),
        ]

        for col, (val, fmt_num) in enumerate(datos, start=1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value         = val
            cell.number_format = fmt_num
            cell.border        = borde
            cell.alignment     = Alignment(
                horizontal="center" if col in cols_centradas else "general",
                vertical="center"
            )

    # ── Anchos de columna ──────────────────────────────────────────────────────
    anchos = [10, 12, 6, 8, 14, 22, 12, 14, 14, 12, 14, 14, 12, 14, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # ── Serializar a bytes ─────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()